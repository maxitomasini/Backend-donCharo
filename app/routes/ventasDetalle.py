from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
from typing import List, Optional
from datetime import datetime, timedelta
import os
from app import models, schemas
from app.database import get_db
from app.auth import require_role, get_current_user
from fastapi.responses import StreamingResponse
from io import BytesIO


router = APIRouter(prefix="/ventas-detalle", tags=["ventas-detalle"])

@router.get("/")
def listar_ventas_detalle(
    skip: int = Query(0, ge=0, description="Número de registros a saltar"),
    limit: int = Query(50, ge=1, le=100, description="Número de registros a retornar"),
    fecha_desde: Optional[str] = Query(None, description="Fecha desde (YYYY-MM-DD)"),
    fecha_hasta: Optional[str] = Query(None, description="Fecha hasta (YYYY-MM-DD)"),
    usuario_id: Optional[int] = Query(None, description="Filtrar por usuario"),
    metodo_pago: Optional[str] = Query(None, description="Filtrar por método de pago"),
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(require_role(["admin", "superadmin"]))
):
    """
    Listar todas las ventas con detalle completo.
    Solo accesible para ADMIN y SUPERADMIN.
    Ordenadas de más reciente a más antigua.
    """
    # Query base con join de usuario
    query = db.query(models.Venta).join(models.Usuario)
    
    # Filtro por rango de fechas
    if fecha_desde:
        try:
            fecha_desde_dt = datetime.strptime(fecha_desde, "%Y-%m-%d")
            query = query.filter(models.Venta.fecha >= fecha_desde_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Formato de fecha_desde inválido. Use YYYY-MM-DD")
    
    if fecha_hasta:
        try:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(models.Venta.fecha < fecha_hasta_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Formato de fecha_hasta inválido. Use YYYY-MM-DD")
    
    # Filtro por usuario
    if usuario_id:
        query = query.filter(models.Venta.usuario_id == usuario_id)
    
    # Filtro por método de pago
    if metodo_pago:
        query = query.filter(models.Venta.metodo_pago == metodo_pago)
    
    # Ordenar de más reciente a más antigua
    query = query.order_by(desc(models.Venta.fecha))
    
    # Obtener total
    total = query.count()
    
    # Aplicar paginación
    ventas = query.offset(skip).limit(limit).all()
    
    # Formatear respuesta con detalles
    ventas_detalle = []
    for venta in ventas:
        # Obtener items de la venta
        items = db.query(models.ItemVenta).filter(
            models.ItemVenta.venta_id == venta.id
        ).all()
        
        items_detalle = []
        for item in items:
            producto = db.query(models.Producto).filter(
                models.Producto.id == item.producto_id
            ).first()
            
            items_detalle.append({
                "producto_id": item.producto_id,
                "producto_nombre": producto.nombre if producto else "Producto eliminado",
                "cantidad": item.cantidad,
                "precio_unitario": float(item.precio_unitario),
                "subtotal": float(item.subtotal)
            })
        
        ventas_detalle.append({
            "id": venta.id,
            "fecha": venta.fecha.isoformat(),
            "total": float(venta.total),
            "metodo_pago": venta.metodo_pago,
            "usuario_id": venta.usuario_id,
            "usuario_nombre": venta.usuario.username,
            "usuario_nombre_completo": venta.usuario.nombre_completo,
            "observaciones": venta.observaciones,
            "items": items_detalle,
            "cantidad_items": len(items_detalle)
        })
    
    return {
        "ventas": ventas_detalle,
        "total": total,
        "skip": skip,
        "limit": limit,
        "has_more": (skip + limit) < total
    }


@router.get("/exportar")
async def exportar_ventas_excel(
    fecha_desde: Optional[str] = Query(None, description="Fecha desde (YYYY-MM-DD)"),
    fecha_hasta: Optional[str] = Query(None, description="Fecha hasta (YYYY-MM-DD)"),
    usuario_id: Optional[int] = Query(None, description="Filtrar por usuario"),
    metodo_pago: Optional[str] = Query(None, description="Filtrar por método de pago"),
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(require_role(["admin", "superadmin"]))
):
    """
    Exportar ventas a Excel.
    Genera un archivo .xlsx con todas las ventas filtradas.
    """
    # Query base
    query = db.query(models.Venta).join(models.Usuario)
    
    # Aplicar filtros (mismo código que listar)
    if fecha_desde:
        try:
            fecha_desde_dt = datetime.strptime(fecha_desde, "%Y-%m-%d")
            query = query.filter(models.Venta.fecha >= fecha_desde_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Formato de fecha_desde inválido")
    
    if fecha_hasta:
        try:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(models.Venta.fecha < fecha_hasta_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Formato de fecha_hasta inválido")
    
    if usuario_id:
        query = query.filter(models.Venta.usuario_id == usuario_id)
    
    if metodo_pago:
        query = query.filter(models.Venta.metodo_pago == metodo_pago)
    
    # Ordenar de más reciente a más antigua
    query = query.order_by(desc(models.Venta.fecha))
    
    # Obtener todas las ventas (sin límite)
    ventas = query.all()
    
    if not ventas:
        raise HTTPException(status_code=404, detail="No hay ventas para exportar con los filtros aplicados")
    
    # Crear archivo Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    
    # ============ HOJA 1: RESUMEN DE VENTAS ============
    ws_resumen = wb.active
    ws_resumen.title = "Resumen Ventas"
    
    # Headers
    headers_resumen = [
        "ID Venta", "Fecha", "Hora", "Usuario", "Nombre Completo", 
        "Método Pago", "Total", "Cantidad Items", "Observaciones"
    ]
    
    # Estilos
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Escribir headers
    for col, header in enumerate(headers_resumen, 1):
        cell = ws_resumen.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Escribir datos
    row = 2
    for venta in ventas:
        items_count = db.query(models.ItemVenta).filter(
            models.ItemVenta.venta_id == venta.id
        ).count()
        
        ws_resumen.cell(row, 1, venta.id)
        ws_resumen.cell(row, 2, venta.fecha.strftime("%Y-%m-%d"))
        ws_resumen.cell(row, 3, venta.fecha.strftime("%H:%M:%S"))
        ws_resumen.cell(row, 4, venta.usuario.username)
        ws_resumen.cell(row, 5, venta.usuario.nombre_completo or "")
        ws_resumen.cell(row, 6, venta.metodo_pago.upper())
        ws_resumen.cell(row, 7, float(venta.total))
        ws_resumen.cell(row, 8, items_count)
        ws_resumen.cell(row, 9, venta.observaciones or "")
        
        # Formato moneda
        ws_resumen.cell(row, 7).number_format = '$#,##0.00'
        
        # Bordes
        for col in range(1, 10):
            ws_resumen.cell(row, col).border = border
            ws_resumen.cell(row, col).alignment = Alignment(vertical='center')
        
        row += 1
    
    # Ajustar anchos de columna
    ws_resumen.column_dimensions['A'].width = 10
    ws_resumen.column_dimensions['B'].width = 12
    ws_resumen.column_dimensions['C'].width = 10
    ws_resumen.column_dimensions['D'].width = 15
    ws_resumen.column_dimensions['E'].width = 20
    ws_resumen.column_dimensions['F'].width = 12
    ws_resumen.column_dimensions['G'].width = 12
    ws_resumen.column_dimensions['H'].width = 12
    ws_resumen.column_dimensions['I'].width = 30
    
    # Fila de totales
    total_row = row
    ws_resumen.cell(total_row, 6, "TOTAL:")
    ws_resumen.cell(total_row, 6).font = Font(bold=True)
    ws_resumen.cell(total_row, 7, f"=SUM(G2:G{row-1})")
    ws_resumen.cell(total_row, 7).font = Font(bold=True)
    ws_resumen.cell(total_row, 7).number_format = '$#,##0.00'
    ws_resumen.cell(total_row, 7).fill = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")
    
    # ============ HOJA 2: DETALLE POR PRODUCTO ============
    ws_detalle = wb.create_sheet("Detalle Productos")
    
    headers_detalle = [
        "ID Venta", "Fecha", "Usuario", "Producto", "Cantidad", 
        "Precio Unitario", "Subtotal", "Método Pago"
    ]
    
    # Headers
    for col, header in enumerate(headers_detalle, 1):
        cell = ws_detalle.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    row = 2
    for venta in ventas:
        items = db.query(models.ItemVenta).filter(
            models.ItemVenta.venta_id == venta.id
        ).all()
        
        for item in items:
            producto = db.query(models.Producto).filter(
                models.Producto.id == item.producto_id
            ).first()
            
            ws_detalle.cell(row, 1, venta.id)
            ws_detalle.cell(row, 2, venta.fecha.strftime("%Y-%m-%d %H:%M"))
            ws_detalle.cell(row, 3, venta.usuario.username)
            ws_detalle.cell(row, 4, producto.nombre if producto else "Producto eliminado")
            ws_detalle.cell(row, 5, item.cantidad)
            ws_detalle.cell(row, 6, float(item.precio_unitario))
            ws_detalle.cell(row, 7, float(item.subtotal))
            ws_detalle.cell(row, 8, venta.metodo_pago.upper())
            
            # Formato
            ws_detalle.cell(row, 6).number_format = '$#,##0.00'
            ws_detalle.cell(row, 7).number_format = '$#,##0.00'
            
            # Bordes
            for col in range(1, 9):
                ws_detalle.cell(row, col).border = border
                ws_detalle.cell(row, col).alignment = Alignment(vertical='center')
            
            row += 1
    
    # Anchos
    ws_detalle.column_dimensions['A'].width = 10
    ws_detalle.column_dimensions['B'].width = 16
    ws_detalle.column_dimensions['C'].width = 15
    ws_detalle.column_dimensions['D'].width = 35
    ws_detalle.column_dimensions['E'].width = 10
    ws_detalle.column_dimensions['F'].width = 14
    ws_detalle.column_dimensions['G'].width = 12
    ws_detalle.column_dimensions['H'].width = 12
    
    # Fila de totales
    total_row_detalle = row
    ws_detalle.cell(total_row_detalle, 6, "TOTAL:")
    ws_detalle.cell(total_row_detalle, 6).font = Font(bold=True)
    ws_detalle.cell(total_row_detalle, 7, f"=SUM(G2:G{row-1})")
    ws_detalle.cell(total_row_detalle, 7).font = Font(bold=True)
    ws_detalle.cell(total_row_detalle, 7).number_format = '$#,##0.00'
    ws_detalle.cell(total_row_detalle, 7).fill = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")
    
    # ============ HOJA 3: ESTADÍSTICAS ============
    ws_stats = wb.create_sheet("Estadísticas")
    
    # Título
    ws_stats.cell(1, 1, "ESTADÍSTICAS DE VENTAS")
    ws_stats.cell(1, 1).font = Font(bold=True, size=14, color="1E40AF")
    ws_stats.merge_cells('A1:B1')
    
    # Período
    fecha_min = min(venta.fecha for venta in ventas)
    fecha_max = max(venta.fecha for venta in ventas)
    
    ws_stats.cell(2, 1, "Período:")
    ws_stats.cell(2, 2, f"{fecha_min.strftime('%Y-%m-%d')} a {fecha_max.strftime('%Y-%m-%d')}")
    ws_stats.cell(2, 1).font = Font(bold=True)
    
    # Estadísticas
    row = 4
    stats = [
        ("Total Ventas:", len(ventas)),
        ("Total Recaudado:", f"=Resumen_Ventas!G{total_row}"),
        ("Promedio por Venta:", f"=B5/{len(ventas)}"),
        ("Ventas Efectivo:", len([v for v in ventas if v.metodo_pago == 'efectivo'])),
        ("Ventas Normal:", len([v for v in ventas if v.metodo_pago == 'normal'])),
    ]
    
    for label, value in stats:
        ws_stats.cell(row, 1, label)
        ws_stats.cell(row, 1).font = Font(bold=True)
        
        if isinstance(value, str) and value.startswith('='):
            ws_stats.cell(row, 2, value)
            if 'G' in value:
                ws_stats.cell(row, 2).number_format = '$#,##0.00'
        else:
            ws_stats.cell(row, 2, value)
            if row in [5, 6]:  # Total y Promedio
                ws_stats.cell(row, 2).number_format = '$#,##0.00'
        
        row += 1
    
    ws_stats.column_dimensions['A'].width = 20
    ws_stats.column_dimensions['B'].width = 20
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"ventas_{timestamp}.xlsx"

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return StreamingResponse(
    file_stream,
    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    headers={
        "Content-Disposition": f'attachment; filename="{filename}"'
    }
)


@router.get("/estadisticas")
def obtener_estadisticas(
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(require_role(["admin", "superadmin"]))
):
    """
    Obtener estadísticas generales de ventas.
    """
    query = db.query(models.Venta)
    
    # Filtros de fecha
    if fecha_desde:
        fecha_desde_dt = datetime.strptime(fecha_desde, "%Y-%m-%d")
        query = query.filter(models.Venta.fecha >= fecha_desde_dt)
    
    if fecha_hasta:
        fecha_hasta_dt = datetime.strptime(fecha_hasta, "%Y-%m-%d") + timedelta(days=1)
        query = query.filter(models.Venta.fecha < fecha_hasta_dt)
    
    # Estadísticas
    total_ventas = query.count()
    total_recaudado = query.with_entities(func.sum(models.Venta.total)).scalar() or 0
    ventas_efectivo = query.filter(models.Venta.metodo_pago == 'efectivo').count()
    ventas_normal = query.filter(models.Venta.metodo_pago == 'normal').count()
    
    promedio_venta = float(total_recaudado) / total_ventas if total_ventas > 0 else 0
    
    return {
        "total_ventas": total_ventas,
        "total_recaudado": float(total_recaudado),
        "promedio_venta": promedio_venta,
        "ventas_efectivo": ventas_efectivo,
        "ventas_normal": ventas_normal,
        "porcentaje_efectivo": (ventas_efectivo / total_ventas * 100) if total_ventas > 0 else 0
    }